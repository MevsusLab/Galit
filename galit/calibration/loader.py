"""CSV/XLSX history loading, validation, templates and WellCase mapping."""
from __future__ import annotations

import csv
from dataclasses import asdict
import hashlib
import json
from pathlib import Path
from typing import Any

from .schema import (HistoryDataset, InputProvenance, LEAKAGE_COLUMNS,
    OPTIONAL_SNAPSHOT_COLUMNS, SCHEMA_VERSION, SNAPSHOT_COLUMNS,
    ValidationResult, WellSnapshot, finite_float, parse_time)

RANGES = {
    "depth_m": (1, 15000), "tubing_id_m": (0.01, 1),
    "inclination_deg": (0, 90), "roughness_m": (0, 0.02),
    "q_oil_m3d": (0, 100000), "q_water_m3d": (0, 100000),
    "gor_m3m3": (0, 100000), "gamma_oil": (0.5, 1.5),
    "gamma_gas": (0.1, 3), "salinity_ppm": (0, 600000),
    "surface_tension_n_m": (0.001, 0.2), "t_surface_c": (-80, 80),
    "geothermal_grad_k_m": (0, 0.2), "k_earth_w_mk": (0.1, 20),
    "alpha_earth_m2_s": (1e-9, 1e-3), "u_to_w_m2k": (0.01, 1000),
    "r_to_m": (0.001, 1), "r_wb_m": (0.001, 5),
    "cp_fluid_j_kgk": (100, 10000), "production_days": (1e-6, 100000),
    "na_mg_l": (0, 500000), "cl_mg_l": (0, 700000),
    "ca_mg_l": (0, 300000), "mg_mg_l": (0, 200000),
    "k_mg_l": (0, 200000), "hco3_mg_l": (0, 100000),
    "so4_mg_l": (0, 200000), "ph": (0, 14), "water_t_c": (-20, 250),
    "water_p_pa": (0, 2e8), "wat_stock_tank_c": (-50, 150),
    "wax_content_pct": (0, 100), "co2_mol_frac": (0, 1),
    "inhibitor_efficiency": (0, 1), "p_wellhead_pa": (0, 2e8),
}
NUMERIC = frozenset(RANGES) | {"target_temperature_c", "target_pressure_pa",
    "measurement_depth_m", "target_wax_onset_m", "target_corrosion_mm_y",
    "event_label", "risk_label"}
VALID_SOURCE = {"measured", "derived", "laboratory"}
VALID_QUALITY = {"good", "questionable", "bad"}


def _charge_balance(row: dict[str, Any]) -> float | None:
    # meq/L using equivalent weights; positive/negative imbalance is diagnostic only.
    c = (row["na_mg_l"]/22.99 + row["k_mg_l"]/39.10 +
         row["ca_mg_l"]/20.04 + row["mg_mg_l"]/12.15)
    a = (row["cl_mg_l"]/35.45 + row["hco3_mg_l"]/61.02 + row["so4_mg_l"]/48.03)
    return None if c + a == 0 else 100 * (c - a) / (c + a)


def validate_rows(rows: list[dict[str, Any]], *, strict: bool = True) -> ValidationResult:
    result = ValidationResult(rows=[])
    seen: set[tuple[str, str]] = set()
    last_by_well: dict[str, Any] = {}
    for index, original in enumerate(rows, 2):
        row = {str(k).strip(): v for k, v in original.items()}
        missing = [c for c in SNAPSHOT_COLUMNS if c not in row or row[c] in (None, "")]
        if missing:
            result.errors.append(f"row {index}: missing required columns/values: {', '.join(missing)}")
            continue
        if str(row["schema_version"]) != SCHEMA_VERSION:
            result.errors.append(f"row {index}: unsupported schema_version {row['schema_version']!r}")
        try:
            timestamp = parse_time(row["timestamp"])
        except ValueError as exc:
            result.errors.append(f"row {index}: {exc}")
            continue
        well_id = str(row["well_id"]).strip()
        if not well_id:
            result.errors.append(f"row {index}: empty well_id")
        key = (well_id, timestamp.isoformat())
        if key in seen:
            result.errors.append(f"row {index}: duplicate (well_id, timestamp) {key}")
        seen.add(key)
        if well_id in last_by_well and timestamp < last_by_well[well_id]:
            result.errors.append(f"row {index}: timestamps are not ordered for well {well_id}")
        last_by_well[well_id] = timestamp
        if row["source"] not in VALID_SOURCE:
            result.errors.append(f"row {index}: source must be one of {sorted(VALID_SOURCE)}")
        if row["quality"] not in VALID_QUALITY:
            result.errors.append(f"row {index}: quality must be one of {sorted(VALID_QUALITY)}")
        parsed = dict(row)
        for name in NUMERIC:
            if name in row and row[name] not in (None, ""):
                try:
                    parsed[name] = finite_float(row[name], name)
                except ValueError as exc:
                    result.errors.append(f"row {index}: {exc}")
        for name, (low, high) in RANGES.items():
            value = parsed.get(name)
            if isinstance(value, (int, float)) and not low <= value <= high:
                result.errors.append(f"row {index}: {name}={value} outside [{low}, {high}]")
        parsed["well_id"], parsed["timestamp"] = well_id, timestamp
        if all(isinstance(parsed.get(k), (float, int)) for k in
               ("na_mg_l", "k_mg_l", "ca_mg_l", "mg_mg_l", "cl_mg_l", "hco3_mg_l", "so4_mg_l")):
            imbalance = _charge_balance(parsed)
            if imbalance is not None and abs(imbalance) > 10:
                result.warnings.append(f"row {index}: ion charge imbalance {imbalance:.1f}% (>10%)")
        result.rows.append(parsed)
    if strict:
        result.raise_for_errors()
    return result


def _csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl") from exc
    ws = load_workbook(path, read_only=True, data_only=True)["snapshots"]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(v) for v in next(iterator)]
    return [dict(zip(headers, values)) for values in iterator if any(v is not None for v in values)]


def load_history(path: str | Path, *, strict: bool = True) -> HistoryDataset:
    path = Path(path)
    if path.suffix.lower() == ".csv": rows = _csv_rows(path)
    elif path.suffix.lower() == ".xlsx": rows = _xlsx_rows(path)
    else: raise ValueError("history must be .csv or .xlsx")
    validated = validate_rows(rows, strict=strict)
    snapshots = [WellSnapshot(r, InputProvenance(str(r["source"]), str(r["quality"]),
        r["timestamp"], str(r["schema_version"]))) for r in validated.rows]
    canonical = json.dumps([{k: (v.isoformat() if hasattr(v, "isoformat") else v)
        for k, v in sorted(s.values.items())} for s in snapshots], sort_keys=True,
        ensure_ascii=False, separators=(",", ":"))
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    synthetic = bool(snapshots) and all(str(s.values.get("is_synthetic", "")).lower() in {"1","true","yes"} for s in snapshots)
    return HistoryDataset(snapshots=snapshots, warnings=validated.warnings,
                          dataset_hash=digest, synthetic=synthetic)


def template_row(*, example: bool = False) -> dict[str, Any]:
    row = {c: "" for c in SNAPSHOT_COLUMNS + OPTIONAL_SNAPSHOT_COLUMNS}
    if not example: return row
    row.update({"schema_version": SCHEMA_VERSION, "well_id": "SYNTHETIC-001",
      "timestamp": "2025-01-01T00:00:00+00:00", "source": "measured", "quality": "good",
      "depth_m": 3000, "tubing_id_m": .062, "inclination_deg": 10, "roughness_m": 4.6e-5,
      "q_oil_m3d": 10, "q_water_m3d": 70, "gor_m3m3": 60, "gamma_oil": .86,
      "gamma_gas": .75, "salinity_ppm": 250000, "surface_tension_n_m": .03,
      "t_surface_c": 8, "geothermal_grad_k_m": .033, "k_earth_w_mk": 2.5,
      "alpha_earth_m2_s": 1e-6, "u_to_w_m2k": 15, "r_to_m": .038, "r_wb_m": .108,
      "cp_fluid_j_kgk": 2100, "production_days": 365, "na_mg_l": 90000,
      "cl_mg_l": 145000, "ca_mg_l": 10000, "mg_mg_l": 1000, "k_mg_l": 1000,
      "hco3_mg_l": 100, "so4_mg_l": 100, "ph": 6, "water_t_c": 40,
      "water_p_pa": 5e6, "wat_stock_tank_c": 34, "wax_content_pct": 6,
      "co2_mol_frac": .01, "inhibitor_efficiency": 0, "lift_type": "ESP",
      "p_wellhead_pa": 1.5e6, "is_synthetic": True})
    return row


def generate_template(path: str | Path, *, example: bool = False) -> Path:
    path = Path(path); headers = list(SNAPSHOT_COLUMNS + OPTIONAL_SNAPSHOT_COLUMNS)
    row = template_row(example=example)
    if path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers); writer.writeheader()
            if example: writer.writerow(row)
    elif path.suffix.lower() == ".xlsx":
        try: from openpyxl import Workbook
        except ImportError as exc: raise RuntimeError("XLSX support requires openpyxl") from exc
        wb = Workbook(); ws = wb.active; ws.title = "snapshots"; ws.append(headers)
        if example: ws.append([row[h] for h in headers])
        for name, cols in (("observations", ["well_id","timestamp","kind","value","unit","source","quality"]),
                           ("events", ["well_id","timestamp","event_type","label","source","quality"]),
                           ("treatments", ["well_id","timestamp","treatment_type","result","source","quality"])):
            sheet=wb.create_sheet(name); sheet.append(cols)
        wb.save(path)
    else: raise ValueError("template must be .csv or .xlsx")
    return path


def assert_no_leakage(features: list[str] | tuple[str, ...] | set[str]) -> None:
    bad = sorted(set(features) & LEAKAGE_COLUMNS)
    post = sorted(f for f in features if f.startswith("post_") or f.startswith("target_"))
    bad = sorted(set(bad + post))
    if bad: raise ValueError("Leakage fields cannot be model features: " + ", ".join(bad))


def snapshot_to_well_case(snapshot: WellSnapshot):
    """Map without defaults; validation guarantees every constructor input."""
    from galit.integrated import DataProvenance, WellCase
    from galit.scale import WaterAnalysis
    from galit.wax import WaxProperties
    from galit.wellbore import FluidProperties, ProductionRate, ThermalParams, WellGeometry
    v = snapshot.values
    sources = {name: snapshot.provenance.source for name in (
      "geometry.depth_m","geometry.tubing_id_m","geometry.inclination_deg","geometry.roughness_m",
      "rate.q_oil_m3d","rate.q_water_m3d","rate.gor_m3m3","fluid.gamma_oil","fluid.gamma_gas",
      "fluid.salinity_ppm","fluid.surface_tension","thermal.t_surface_c","thermal.geothermal_grad",
      "thermal.k_earth","thermal.alpha_earth","thermal.u_to","thermal.r_to","thermal.r_wb",
      "thermal.cp_fluid","thermal.production_days","water.ions_mg_l","water.ph","water.t_c",
      "water.p_pa","wax.wat_stock_tank_c","wax.wax_content_pct","co2_mol_frac",
      "inhibitor_efficiency","lift_type","p_wellhead_pa")}
    return WellCase(name=str(v["well_id"]),
      geometry=WellGeometry(v["depth_m"],v["tubing_id_m"],v["inclination_deg"],v["roughness_m"]),
      rate=ProductionRate(v["q_oil_m3d"],v["q_water_m3d"],v["gor_m3m3"]),
      fluid=FluidProperties(v["gamma_oil"],v["gamma_gas"],v["salinity_ppm"],v["surface_tension_n_m"]),
      thermal=ThermalParams(v["t_surface_c"],v["geothermal_grad_k_m"],v["k_earth_w_mk"],v["alpha_earth_m2_s"],v["u_to_w_m2k"],v["r_to_m"],v["r_wb_m"],v["cp_fluid_j_kgk"],v["production_days"]),
      water=WaterAnalysis({"Na":v["na_mg_l"],"Cl":v["cl_mg_l"],"Ca":v["ca_mg_l"],"Mg":v["mg_mg_l"],"K":v["k_mg_l"],"HCO3":v["hco3_mg_l"],"SO4":v["so4_mg_l"]},v["ph"],v["water_t_c"],v["water_p_pa"]),
      wax=WaxProperties(v["wat_stock_tank_c"],v["wax_content_pct"]),
      co2_mol_frac=v["co2_mol_frac"], inhibitor_efficiency=v["inhibitor_efficiency"],
      lift_type=str(v["lift_type"]), p_wellhead_pa=v["p_wellhead_pa"],
      provenance=DataProvenance(sources=sources))
